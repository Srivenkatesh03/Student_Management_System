from django.shortcuts import redirect, render, get_object_or_404
import csv
from django.http import HttpResponse
from django.contrib import messages
from core.forms import StudentForm, MarksForm, AttendanceForm
from .models import Student, Attendance
from django.db.models import Q
from django.core.paginator import Paginator
from django.contrib.auth.decorators import login_required,user_passes_test
from django.db.models import Count
from django.utils.timezone import now



def is_staff_user(user):
    return user.is_authenticated and user.is_staff

# Create your views here.
def home(request):
    return render(request, 'core/home.html')

@login_required
def student_list(request):
    query = request.GET.get('q','')
    students = Student.objects.all().order_by('-created_at')
    if query:
        students = students.filter(
        Q(name__icontains=query) |
        Q(email__icontains=query) |
        Q(department__icontains=query) |
        Q(roll_number__icontains=query) 
        )
    paginator = Paginator(students,5)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'core/student_list.html', {'page_obj':page_obj,'query':query})

@login_required
@user_passes_test(is_staff_user)
def student_create(request):
    if request.method == 'POST':
        form = StudentForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Student added successfully.')
            return redirect('student_list')
    else:
        form = StudentForm()

    return render(request, 'core/student_form.html', {'form': form})

@login_required
@user_passes_test(is_staff_user)
def student_update(request, pk):
    student = get_object_or_404(Student, pk=pk)

    if request.method == 'POST':
        form = StudentForm(request.POST, request.FILES, instance=student)
        if form.is_valid():
            form.save()
            messages.success(request, 'Student updated successfully.')
            return redirect('student_list')
    else:
        form = StudentForm(instance=student)

    return render(request, 'core/student_form.html', {'form': form})

@login_required
@user_passes_test(is_staff_user)
def student_delete(request, pk):
    student = get_object_or_404(Student, pk=pk)

    if request.method == 'POST':
        student.delete()
        messages.success(request,'Student deleted succesfully.')
        return redirect('student_list')
    
    return render(request,'core/student_confirm_delete.html',{'student':student})

@login_required
def student_detail(request, pk):
    student = get_object_or_404(Student,pk=pk)
    return render(request, 'core/student_detail.html', {'student':student})


@login_required
@user_passes_test(is_staff_user)
def marks_create(request):
    if request.method == 'POST':
        form = MarksForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('student_list')
    else:
        form = MarksForm()
    return render(request, 'core/marks_form.html',{'form':form})

def student_report(request, pk):
    student = get_object_or_404(Student, pk=pk)
    marks = student.marks.select_related('subject')

    context = {
        'student': student,
        'marks': marks,
    }
    return render(request, 'core/student_report.html', context)

@login_required
@user_passes_test(is_staff_user)
def export_student_details_csv(request,pk):
    student= get_object_or_404(Student,pk=pk)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="student_{student.id}details.csv"'

    writer = csv.writer(response)
    
    writer.writerow({
        'Name','Email',"Department",'Year','Total Marks', 'Percentage', 'Result'
    })

    writer.writerow([
        student.name,
        student.email,
        student.get_department_display(),
        student.year,
        student.total_marks_obtained(),
        student.percentage(),
        student.result(),
    ])

    return response


@login_required
@user_passes_test(is_staff_user)
def export_student_marks_csv(request, pk):
    student = get_object_or_404(Student, pk=pk)
    marks = student.marks.select_related('subject')

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = (
        f'attachment; filename="student_{student.id}_marks.csv"'
    )

    writer = csv.writer(response)

    # Header
    writer.writerow([
        'Subject',
        'Marks Obtained',
        'Max Marks'
    ])

    # Rows
    for mark in marks:
        writer.writerow([
            mark.subject.name,
            mark.marks_obtained,
            mark.subject.max_marks
        ])

    return response

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required, user_passes_test
from .models import Student


def is_staff_user(user):
    return user.is_authenticated and user.is_staff


@login_required
@user_passes_test(is_staff_user)
def export_all_students_excel(request):
    wb = Workbook()

    students_sheet = wb.active
    students_sheet.title = "Students"

    student_headers = [
        "Student ID", "Name", "Email", "Department",
        "Year", "Total Marks", "Percentage", "Result"
    ]
    students_sheet.append(student_headers)

    for student in Student.objects.all():
        students_sheet.append([
            student.id,
            student.name,
            student.email,
            student.get_department_display(),
            student.year,
            student.total_marks_obtained(),
            student.percentage(),
            student.result(),
        ])

    for col in range(1, len(student_headers) + 1):
        students_sheet.column_dimensions[get_column_letter(col)].width = 20

    marks_sheet = wb.create_sheet(title="Marks")

    marks_headers = [
        "Student ID", "Student Name",
        "Subject", "Marks Obtained", "Max Marks"
    ]
    marks_sheet.append(marks_headers)

    for student in Student.objects.all():
        for mark in student.marks.select_related("subject"):
            marks_sheet.append([
                student.id,
                student.name,
                mark.subject.name,
                mark.marks_obtained,
                mark.subject.max_marks,
            ])

    for col in range(1, len(marks_headers) + 1):
        marks_sheet.column_dimensions[get_column_letter(col)].width = 20

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="All_Students_Report.xlsx"'

    wb.save(response)
    return response

@login_required
@user_passes_test(is_staff_user)
def dashboard(request):
    students = Student.objects.all()

    total_students = students.count()

    passed_students = sum(
        1 for student in students if student.result() == "PASS"
    )
    failed_students = total_students - passed_students

    department_stats = (
        students.values("department")
        .annotate(count=Count("id"))
    )

    percentages = [
        student.percentage() for student in students
        if student.total_max_marks() > 0
    ]
    avg_percentage = round(
        sum(percentages) / len(percentages), 2
    ) if percentages else 0

    context = {
        "total_students": total_students,
        "passed_students": passed_students,
        "failed_students": failed_students,
        "department_stats": department_stats,
        "avg_percentage": avg_percentage,
    }

    return render(request, "core/dashboard.html", context)

@login_required
@user_passes_test(is_staff_user)
def mark_attendance(request):
    date = request.GET.get('date') or now().date()
    students = Student.objects.all()
    if request.method == "POST":
        for student in students:
            status = request.POST.get(f"status_{student.id}")
            if status:
                Attendance.objects.update_or_create(
                    student=student,
                    date=date,
                    defaults={"status":status}
                ) 
        return redirect("dashboard")
    
    context = {
        "students":students,
        "date":date
    }
    return render(request,"core/attendance_form.html",context)

